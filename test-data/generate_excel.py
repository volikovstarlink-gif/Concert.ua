import json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

with open('test-data/tests_extracted.json', 'r', encoding='utf-8') as f:
    tests = json.load(f)

# Ukrainian translations for section names
SECTION_UA = {
    'Доступність / a11y (Accessibility)': 'Доступність (a11y)',
    'Особистий кабінет (Personal Account)': 'Особистий кабінет',
    'Кошик та оформлення (Cart & Purchase Flow)': 'Кошик та оформлення',
    'Каталог подій (Event Catalog)': 'Каталог подій',
    'Каталог — фільтри та сортування (Catalog Filters & Sorting)': 'Каталог — фільтри та сортування',
    'Контентні сторінки (Content & Info Pages)': 'Контентні сторінки',
    'Cookie банер та Футер (Cookie Banner & Footer)': 'Cookie банер та Футер',
    'Картка події — деталі (Event Page Details)': 'Картка події — деталі',
    'Картка події (Event Detail Page)': 'Картка події',
    'Головна сторінка (Home Page)': 'Головна сторінка',
    'Головна сторінка — секції (Home Page Sections)': 'Головна сторінка — секції',
    'Локалізація / і18n (Localization)': 'Локалізація / i18n',
    'Навігація (Navigation & Header)': 'Навігація та заголовок',
    'Оплата (Payment)': 'Оплата',
    'Адаптивність / Мобайл (Responsive / Mobile)': 'Адаптивність / Мобайл',
    'Пошук (Search)': 'Пошук',
}

# Ukrainian test title translations
TITLE_UA = {
    'Page has a valid lang attribute': 'Сторінка має валідний атрибут lang',
    'Page has a main heading (h1)': 'Сторінка має основний заголовок (h1)',
    'Images have alt attributes': 'Зображення мають атрибути alt',
    'Page title is not empty': 'Заголовок сторінки не порожній',
    'Links have discernible text': 'Посилання мають зрозумілий текст',
    'Form inputs have associated labels or placeholders': 'Поля введення мають мітки або placeholder',
    'Page is navigable with keyboard (Tab key)': 'Сторінка доступна з клавіатури (Tab)',
    'Color contrast — text is readable': 'Контрастність кольорів — текст читабельний',
    'Main landmark role exists': 'Є основний landmark-елемент (main)',
    'Buttons have proper role': 'Кнопки мають правильну роль (button)',
    'Profile page loads successfully': 'Сторінка профілю завантажується успішно',
    'Login page loads without server error': 'Сторінка входу завантажується без серверної помилки',
    'Profile page has valid title': 'Сторінка профілю має валідний title',
    'Tickets page loads from profile': 'Сторінка квитків завантажується з профілю',
    'Certificates page loads from profile': 'Сторінка сертифікатів завантажується з профілю',
    'Preferences/notifications page loads': 'Сторінка сповіщень/налаштувань завантажується',
    'Wishlist page loads': 'Сторінка "Обране" завантажується',
    'Profile link accessible from header': 'Посилання на профіль доступне у заголовку',
    'Profile page loads with valid HTTP status': 'Сторінка профілю повертає коректний HTTP-статус',
    'Profile page is accessible via keyboard navigation': 'Сторінка профілю доступна з клавіатури',
    'News/updates page loads': 'Сторінка новин завантажується',
    'Profile page uses HTTPS': 'Сторінка профілю використовує HTTPS',
    'Event page has buy/date-selection button': 'Сторінка події має кнопку "Обрати дату і час"',
    'Profile tickets page loads for authenticated user': 'Сторінка квитків завантажується для авторизованого користувача',
    'Profile tickets page has valid title': 'Сторінка квитків має валідний title',
    'Event buy button is clickable': 'Кнопка купівлі квитків клікабельна',
    'Profile certificates page loads': 'Сторінка сертифікатів завантажується',
    'Header profile section is visible': 'Секція профілю у заголовку видима',
    'Gift certificate event page loads': 'Сторінка подарункового сертифікату завантажується',
    'Event page maintains state after reload': 'Сторінка події зберігає стан після перезавантаження',
    'Catalog page loads and displays events': 'Сторінка каталогу завантажується та показує події',
    'Category page title (h1) is displayed': 'Заголовок категорії (h1) відображається',
    'Concert category page loads correctly': 'Сторінка категорії "Концерти" завантажується коректно',
    'Theater category page loads correctly': 'Сторінка категорії "Театр" завантажується коректно',
    'Event card displays content and is an anchor link': 'Картка події відображає контент і є посиланням',
    'Clicking event card navigates to event detail page': 'Клік на картку події переходить на сторінку деталей',
    'Multiple event cards are displayed in catalog': 'У каталозі відображається кілька карток подій',
    'Breadcrumb navigation is present on catalog page': 'Хлібні крихти присутні на сторінці каталогу',
    'Standup/humor category loads': 'Категорія "Стендап" завантажується',
    'Festival category loads': 'Категорія "Фестивалі" завантажується',
    'Catalog page has valid meta title': 'Сторінка каталогу має валідний meta title',
    'Genre sub-filter tabs are displayed (Поп, Рок, Класика, etc.)': 'Підфільтри жанрів відображаються (Поп, Рок, Класика тощо)',
    '"Всі" genre tab links to main concerts page': 'Вкладка "Всі" веде на головну сторінку концертів',
    'Clicking "Рок" genre tab navigates to rock sub-page': 'Клік на "Рок" переходить на підсторінку року',
    'Rock sub-genre page has correct h1': 'Підсторінка жанру "Рок" має коректний h1',
    'Filter button is visible on catalog page': 'Кнопка "Фільтри" видима на сторінці каталогу',
    'Sort dropdown is present on catalog page': 'Випадаючий список сортування присутній',
    'Event card shows price information': 'Картка події показує ціну (₴)',
    'Event card shows date information': 'Картка події показує дату',
    'Event card shows venue/city information': 'Картка події показує місце/місто',
    'Event card has image': 'Картка події має зображення',
    'Theater catalog loads with events': 'Каталог "Театр" завантажується з подіями',
    'Kids catalog loads with events': 'Каталог "Дітям" завантажується з подіями',
    'Venue list page loads with title': 'Сторінка списку майданчиків завантажується',
    'Artist list page loads with title': 'Сторінка списку артистів завантажується',
    'Archive page loads with title': 'Сторінка архіву завантажується',
    'About page loads with content': 'Сторінка "Про нас" завантажується з контентом',
    'Team page loads': 'Сторінка "Команда" завантажується',
    'Ticket offices page loads': 'Сторінка квиткових офісів завантажується',
    'Delivery page loads': 'Сторінка доставки завантажується',
    'Venue detail page (Origin Stage) loads with events': 'Сторінка майданчика (Origin Stage) завантажується з подіями',
    'Venue detail page shows event listings': 'Сторінка майданчика показує список подій',
    'Metacard collection page loads': 'Сторінка метакартки (колекція) завантажується',
    'Metacard page contains event cards': 'Сторінка метакартки містить картки подій',
    'Charity page loads': 'Сторінка благодійності завантажується',
    'Offer/terms page has content': 'Сторінка оферти має контент',
    'FAQ page has questions and answers': 'Сторінка FAQ має запитання та відповіді',
    'Cookie banner appears on first visit': 'Cookie-банер з\'являється при першому відвідуванні',
    'Cookie banner has "Прийняти все" button': 'Cookie-банер має кнопку "Прийняти все"',
    'Cookie banner has "Відхилити" button': 'Cookie-банер має кнопку "Відхилити"',
    'Cookie banner has cookie type checkboxes': 'Cookie-банер має чекбокси типів cookie',
    'Clicking "Прийняти все" hides cookie banner': 'Клік "Прийняти все" приховує cookie-банер',
    'Page contains social media links (Facebook, Instagram, etc.)': 'Сторінка містить посилання на соцмережі (Facebook, Instagram тощо)',
    'Page contains phone number links': 'Сторінка містить посилання з номерами телефону',
    'Footer shows copyright text': 'Футер показує текст копірайту',
    'Footer contains payment logos (Visa, Mastercard)': 'Футер містить логотипи платіжних систем (Visa, Mastercard)',
    'Footer "Архів" link works': 'Посилання "Архів" у футері працює',
    'Footer "Список майданчиків" link works': 'Посилання "Список майданчиків" у футері працює',
    'Footer "Список артистів" link works': 'Посилання "Список артистів" у футері працює',
    'Artist section is displayed with performer name': 'Секція "Артисти" відображається з іменем виконавця',
    'Artist section contains performer image': 'Секція "Артисти" містить зображення виконавця',
    'Reviews section "Враження глядачів" exists': 'Секція "Враження глядачів" існує',
    '"Оцінити" button is present in reviews section': 'Кнопка "Оцінити" присутня у секції відгуків',
    'Related events section "Також може зацікавити" exists': 'Секція "Також може зацікавити" існує',
    'Related events section contains event cards': 'Секція схожих подій містить картки подій',
    'Like button is present on event page': 'Кнопка "Подобається" присутня на сторінці події',
    'Share "Поділитись" button is present': 'Кнопка "Поділитись" присутня',
    '"БІЛЬШЕ ІНФО" description expand link exists': 'Посилання "БІЛЬШЕ ІНФО" для розгортання опису існує',
    'Sticky buy button "ОБРАТИ ДАТУ І ЧАС" is at page bottom': 'Фіксована кнопка "ОБРАТИ ДАТУ І ЧАС" знизу сторінки',
    'Event page shows venue name and address': 'Сторінка події показує назву та адресу майданчика',
    'Event page has category tags (Рок, Поп, Концерти)': 'Сторінка події має теги категорій (Рок, Поп, Концерти)',
    'Spotify embed widget loads on event page': 'Віджет Spotify завантажується на сторінці події',
    'Event with single date loads correctly (/uk/event/ path)': 'Подія з однією датою завантажується коректно (/uk/event/)',
    'Event page displays event title (h1)': 'Сторінка події відображає заголовок (h1)',
    'Event page has event-title element in DOM': 'Сторінка події має елемент event-title у DOM',
    'Event page displays date information': 'Сторінка події відображає інформацію про дату',
    'Event page displays event image': 'Сторінка події відображає зображення',
    'Buy/date selection button is present': 'Кнопка вибору дати/купівлі присутня',
    'Event title text is non-empty': 'Текст заголовку події не порожній',
    'Event page has valid meta title': 'Сторінка події має валідний meta title',
    'Event page URL matches expected pattern': 'URL сторінки події відповідає очікуваному шаблону',
    'Event image has valid src attribute': 'Зображення події має валідний атрибут src',
    'Share button exists on event page': 'Кнопка "Поділитись" існує на сторінці події',
    'Single-date event page loads correctly': 'Сторінка події з однією датою завантажується коректно',
    'Event page loads without server errors': 'Сторінка події завантажується без серверних помилок',
    'Home page loads successfully with HTTP 200': 'Головна сторінка завантажується успішно (HTTP 200)',
    'Logo is visible on the home page': 'Логотип видимий на головній сторінці',
    'Main navigation menu is displayed': 'Головне навігаційне меню відображається',
    'Search icon is visible in header': 'Іконка пошуку видима у заголовку',
    'Main banner/slider is displayed': 'Головний банер/слайдер відображається',
    'Event cards are displayed on the home page': 'Картки подій відображаються на головній сторінці',
    'Footer section is present': 'Секція футера присутня',
    'Profile/account button is accessible': 'Кнопка профілю/акаунту доступна',
    'Clicking logo navigates to home page': 'Клік на логотип переходить на головну сторінку',
    'Page title contains relevant text': 'Заголовок сторінки містить релевантний текст',
    'Event cards are clickable and navigate to event page': 'Картки подій клікабельні та ведуть на сторінку події',
    'Header section is present at top of page': 'Секція заголовку присутня вгорі сторінки',
    'Category navigation links are present in menu': 'Посилання категорій присутні у навігаційному меню',
    'Hero slider has navigation dots': 'Герой-слайдер має навігаційні точки',
    'Hero slider dots are clickable': 'Навігаційні точки слайдера клікабельні',
    '"Топ подій у Києві" section is displayed': 'Секція "Топ подій у Києві" відображається',
    '"Топ подій" section contains event cards': 'Секція "Топ подій" містить картки подій',
    'Date filter "Сьогодні" button is present': 'Кнопка фільтру "Сьогодні" присутня',
    'Date filter "Завтра" button is present': 'Кнопка фільтру "Завтра" присутня',
    'Date filter "Цього вікенду" button is present': 'Кнопка фільтру "Цього вікенду" присутня',
    '"Обрати період" date picker button is present': 'Кнопка "Обрати період" присутня',
    'Category radio tabs are displayed (Концерти, Театр, etc.)': 'Радіо-вкладки категорій відображаються (Концерти, Театр тощо)',
    '"Рекомендації для вас" section is displayed': 'Секція "Рекомендації для вас" відображається',
    'Recommendation cards contain event info (image, title, price)': 'Картки рекомендацій містять інфо (зображення, назва, ціна)',
    '"Нове у продажу" section exists': 'Секція "Нове у продажу" існує',
    '"Популярне у Києві" metacard section exists': 'Секція "Популярне у Києві" існує',
    'Venue sections display on home page (Origin Stage, МЦКМ)': 'Секції майданчиків відображаються (Origin Stage, МЦКМ)',
    '"Всі події" button is present': 'Кнопка "Всі події" присутня',
    'Carousel scroll buttons exist on home page': 'Кнопки прокрутки каруселей існують на головній',
    'Default page language is Ukrainian': 'Мова сторінки за замовчуванням — українська',
    'Page content contains Ukrainian text': 'Контент сторінки містить українські тексти',
    'Language switcher links are present': 'Посилання перемикача мови присутні',
    'Ukrainian currency format (₴/грн) used for prices': 'Українська валюта (₴/грн) використовується для цін',
    'Date format follows Ukrainian conventions': 'Формат дати відповідає українським конвенціям',
    'UTF-8 encoding is declared': 'Оголошено кодування UTF-8',
    'UI buttons have Ukrainian labels': 'Кнопки інтерфейсу мають українські написи',
    'City names display in Ukrainian': 'Назви міст відображаються українською',
    'Hamburger menu button opens side menu': 'Кнопка гамбургер-меню відкриває бокове меню',
    'Side menu contains category links': 'Бокове меню містить посилання категорій',
    'Nav link "Концерти" navigates to concerts catalog': 'Посилання "Концерти" переходить на каталог концертів',
    'Nav link "Театр" navigates to theater catalog': 'Посилання "Театр" переходить на каталог театру',
    'Nav link "Стендап" navigates to humor catalog': 'Посилання "Стендап" переходить на каталог стендапу',
    'Nav link "Дітям" navigates to kids catalog': 'Посилання "Дітям" переходить на каталог для дітей',
    'Nav link "Танці" navigates to dance catalog': 'Посилання "Танці" переходить на каталог танцю',
    'City selector shows city name "Київ"': 'Селектор міста показує "Київ"',
    'Language switch links exist (uk/en)': 'Посилання перемикання мови існують (uk/en)',
    'English language link navigates to EN version': 'Посилання на англійську версію веде на /en/',
    'Breadcrumbs visible on catalog page': 'Хлібні крихти видимі на сторінці каталогу',
    'Breadcrumbs link navigates back to parent': 'Посилання хлібних крихт повертає на батьківську сторінку',
    'Header remains visible on scroll': 'Заголовок залишається видимим при прокрутці',
    'Support chat button is present': 'Кнопка підтримки (чат) присутня',
    'Event page loads over HTTPS': 'Сторінка події завантажується по HTTPS',
    'Buy button present on event page for active event': 'Кнопка купівлі присутня на сторінці активної події',
    'Event page displays price info': 'Сторінка події відображає інформацію про ціну',
    'Payment info page loads': 'Інформаційна сторінка оплати завантажується',
    'Refund policy page loads': 'Сторінка повернення завантажується',
    'Offer/terms page loads': 'Сторінка оферти завантажується',
    'FAQ page loads': 'Сторінка FAQ завантажується',
    'Delivery info page loads': 'Сторінка інформації про доставку завантажується',
    'Home page renders on mobile viewport (375px)': 'Головна сторінка рендериться на мобільному (375px)',
    'Home page renders on tablet viewport (768px)': 'Головна сторінка рендериться на планшеті (768px)',
    'Hamburger menu button visible on mobile': 'Кнопка гамбургер-меню видима на мобільному',
    'No horizontal scroll on mobile viewport': 'Немає горизонтальної прокрутки на мобільному',
    'Header logo visible on mobile': 'Логотип у заголовку видимий на мобільному',
    'Event cards are displayed on mobile': 'Картки подій відображаються на мобільному',
    'Footer is visible on mobile after scrolling': 'Футер видимий на мобільному після прокрутки',
    'Page content adjusts between viewport sizes': 'Контент сторінки адаптується до розміру вікна',
    'Touch-friendly tap targets on mobile': 'Зручні тап-зони для дотику на мобільному',
    'Viewport meta tag is present': 'Мета-тег viewport присутній',
    'Search icon is visible in the header': 'Іконка пошуку видима у заголовку',
    'Search container is present in header': 'Контейнер пошуку присутній у заголовку',
    'Search input exists with correct attributes': 'Поле пошуку існує з коректними атрибутами',
    'Clicking search icon activates search input': 'Клік на іконку пошуку активує поле введення',
    'Search input accepts Ukrainian characters': 'Поле пошуку приймає українські символи',
    'Search input can be cleared': 'Поле пошуку можна очистити',
    'Search input handles special characters': 'Поле пошуку обробляє спеціальні символи',
    'Multiple search queries work sequentially': 'Декілька пошукових запитів працюють послідовно',
}

# Steps and expected results derived from test purpose
def get_steps_and_expected(test_id, title):
    steps = ''
    expected = ''
    t = title.lower()

    if 'loads' in t or 'завантаж' in t:
        steps = 'Відкрити відповідну сторінку'
        expected = 'Сторінка завантажується без помилок, HTTP 200'
    elif 'visible' in t or 'displayed' in t or 'present' in t or 'exists' in t or 'видим' in t or 'присут' in t or 'відображ' in t:
        steps = 'Відкрити сторінку та знайти елемент'
        expected = 'Елемент видимий на сторінці'
    elif 'click' in t or 'navigat' in t or 'клік' in t or 'перехо' in t:
        steps = 'Відкрити сторінку та клікнути на елемент'
        expected = 'Відбувається перехід на очікувану сторінку'
    elif 'contains' in t or 'містить' in t or 'shows' in t or 'показує' in t:
        steps = 'Відкрити сторінку та перевірити контент'
        expected = 'Очікуваний контент присутній на сторінці'
    elif 'accept' in t or 'input' in t or 'приймає' in t:
        steps = 'Ввести дані в поле введення'
        expected = 'Дані успішно введені та оброблені'
    elif 'mobile' in t or 'мобільн' in t or 'viewport' in t:
        steps = 'Змінити розмір вікна та перевірити елемент'
        expected = 'Елемент коректно відображається на заданому розмірі'
    elif 'https' in t or 'HTTP' in t:
        steps = 'Відкрити сторінку та перевірити протокол'
        expected = 'Використовується HTTPS / HTTP 200'
    else:
        steps = 'Відкрити сторінку та виконати перевірку'
        expected = 'Перевірка пройдена успішно'

    return steps, expected

def get_priority(section, title):
    t = title.lower()
    if any(x in t for x in ['loads', 'http 200', 'завантаж', 'logo', 'логотип', 'buy', 'купів', 'title']):
        return 'Високий'
    if any(x in t for x in ['click', 'navigat', 'клік', 'перехо', 'search', 'пошук', 'filter', 'фільтр']):
        return 'Середній'
    return 'Низький'

wb = Workbook()

# --- CHECKLIST SHEET ---
ws = wb.active
ws.title = 'QA Чеклист'

BRAND_RED = 'E31E24'
header_fill = PatternFill('solid', fgColor=BRAND_RED)
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
data_font = Font(name='Arial', size=10)
border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='top')

headers = ['№', 'ID тесту', 'Розділ', 'Назва тесту', 'Кроки', 'Очікуваний результат', 'Статус', 'Коментар', 'Пріоритет']
col_widths = [5, 13, 30, 55, 35, 40, 15, 25, 12]

for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:I{len(tests)+1}'

# Status dropdown
status_dv = DataValidation(type='list', formula1='"Пройдено ✓,Провалено ✗,Пропущено —,В процесі"', allow_blank=True)
status_dv.error = 'Оберіть значення зі списку'
status_dv.errorTitle = 'Невірне значення'
ws.add_data_validation(status_dv)

# Priority dropdown
prio_dv = DataValidation(type='list', formula1='"Високий,Середній,Низький"', allow_blank=True)
ws.add_data_validation(prio_dv)

# Green/red/yellow conditional colors for status
pass_fill = PatternFill('solid', fgColor='C6EFCE')
fail_fill = PatternFill('solid', fgColor='FFC7CE')
skip_fill = PatternFill('solid', fgColor='FFEB9C')

for i, t in enumerate(tests, 1):
    row = i + 1
    section_ua = SECTION_UA.get(t['section'], t['section'])
    title_ua = TITLE_UA.get(t['title'], t['title'])
    steps, expected = get_steps_and_expected(t['id'], t['title'])
    priority = get_priority(section_ua, t['title'])

    values = [i, t['id'], section_ua, title_ua, steps, expected, '', '', priority]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = data_font
        cell.border = border
        cell.alignment = wrap_align if col_idx in (3, 4, 5, 6) else center_align if col_idx in (1, 2, 7, 9) else wrap_align

    status_dv.add(ws.cell(row=row, column=7))
    prio_dv.add(ws.cell(row=row, column=9))

    # Alternating row background
    if i % 2 == 0:
        for col_idx in range(1, 10):
            ws.cell(row=row, column=col_idx).fill = PatternFill('solid', fgColor='F9F9F9')

# --- SUMMARY SHEET ---
ws2 = wb.create_sheet('Зведення')

# Count per section
from collections import Counter
section_counts = Counter(SECTION_UA.get(t['section'], t['section']) for t in tests)

ws2.cell(row=1, column=1, value='QA ЧЕКЛИСТ — concert.ua').font = Font(name='Arial', bold=True, size=14, color=BRAND_RED)
ws2.merge_cells('A1:F1')

summary_headers = ['Розділ', 'Всього', 'Пройдено ✓', 'Провалено ✗', 'Пропущено —', 'Прогрес']
for col_idx, h in enumerate(summary_headers, 1):
    cell = ws2.cell(row=3, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

ws2.column_dimensions['A'].width = 35
for c in 'BCDEF':
    ws2.column_dimensions[c].width = 15

row = 4
for section, count in sorted(section_counts.items()):
    ws2.cell(row=row, column=1, value=section).font = data_font
    ws2.cell(row=row, column=1).border = border
    ws2.cell(row=row, column=2, value=count).font = data_font
    ws2.cell(row=row, column=2).alignment = center_align
    ws2.cell(row=row, column=2).border = border
    # COUNTIF formulas referencing the main sheet
    col_g_range = f"'QA Чеклист'!G2:G{len(tests)+1}"
    col_c_range = f"'QA Чеклист'!C2:C{len(tests)+1}"
    ws2.cell(row=row, column=3, value=f'=COUNTIFS({col_c_range},A{row},{col_g_range},"Пройдено ✓")').border = border
    ws2.cell(row=row, column=3).alignment = center_align
    ws2.cell(row=row, column=4, value=f'=COUNTIFS({col_c_range},A{row},{col_g_range},"Провалено ✗")').border = border
    ws2.cell(row=row, column=4).alignment = center_align
    ws2.cell(row=row, column=5, value=f'=COUNTIFS({col_c_range},A{row},{col_g_range},"Пропущено —")').border = border
    ws2.cell(row=row, column=5).alignment = center_align
    ws2.cell(row=row, column=6, value=f'=IF(B{row}=0,"-",TEXT(C{row}/B{row},"0%"))').border = border
    ws2.cell(row=row, column=6).alignment = center_align
    row += 1

# Total row
total_row = row
ws2.cell(row=total_row, column=1, value='РАЗОМ').font = Font(name='Arial', bold=True, size=11)
ws2.cell(row=total_row, column=1).border = border
ws2.cell(row=total_row, column=2, value=f'=SUM(B4:B{total_row-1})').font = Font(name='Arial', bold=True)
ws2.cell(row=total_row, column=2).border = border
ws2.cell(row=total_row, column=2).alignment = center_align
ws2.cell(row=total_row, column=3, value=f'=SUM(C4:C{total_row-1})').font = Font(name='Arial', bold=True)
ws2.cell(row=total_row, column=3).border = border
ws2.cell(row=total_row, column=3).alignment = center_align
ws2.cell(row=total_row, column=4, value=f'=SUM(D4:D{total_row-1})').font = Font(name='Arial', bold=True)
ws2.cell(row=total_row, column=4).border = border
ws2.cell(row=total_row, column=4).alignment = center_align
ws2.cell(row=total_row, column=5, value=f'=SUM(E4:E{total_row-1})').font = Font(name='Arial', bold=True)
ws2.cell(row=total_row, column=5).border = border
ws2.cell(row=total_row, column=5).alignment = center_align
ws2.cell(row=total_row, column=6, value=f'=IF(B{total_row}=0,"-",TEXT(C{total_row}/B{total_row},"0%"))').font = Font(name='Arial', bold=True)
ws2.cell(row=total_row, column=6).border = border
ws2.cell(row=total_row, column=6).alignment = center_align

ws2.freeze_panes = 'A4'

output_path = 'test-data/Concert_UA_QA_Checklist.xlsx'
wb.save(output_path)
print(f'Saved {output_path} with {len(tests)} tests')
